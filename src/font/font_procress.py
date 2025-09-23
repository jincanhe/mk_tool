# !/usr/bin/env python
#  -*- coding: utf-8 -*-
# Author: chl
# Email: haolin.chen1991@gmail.com
# 2019.03.11 00:53
# 注意：
# 翻译字典只会越来越大，因此后面要备留反向抓取模式，即找些翻译表哪些key是已经不使用了

import hashlib
import os
import shutil
import time as tm
import sys
import argparse
import logging
from fontTools import subset
import openpyxl as xl
import os.path as path
import io
import json
import re
from os.path import join

logging.basicConfig()

dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(join(dir, "../../"))
sys.path.append(join(dir, "../"))

tempHtml = path.join(dir, 'template/font.html');
tempCss = path.join(dir, 'template/font.css');

nowLDT = tm.localtime()

skipFiles = set([
    "LanguageText.ts",#前段配置
    '.DS_Store'
])

includeFilesRegular = set([
    '.ts',
    '.prefab'
])

def log(content):
    print(content)

def checkIsUTF8(filePath):
    return True

def seg_char(sent, charDict):
    # 首先分割 英文 以及英文和标点
    # Python 2和Python 3兼容处理
    try:
        # 对于Python 2，如果sent是str类型，需要解码为unicode
        if isinstance(sent, str) and sys.version_info[0] < 3:
            sent = sent.decode('utf-8')
    except:
        # 如果出错，可能已经是unicode或Python 3的str
        pass

    pattern_char_1 = re.compile(r'([\W])')
    parts = pattern_char_1.split(sent)
    parts = [p for p in parts if len(p.strip()) > 0]
    
    # 分割中文
    pattern = re.compile(r'([\u4e00-\u9fa5])')
    chars = pattern.split(sent)
    for w in chars:
        for c in w.strip():
            if c not in charDict:
                print("加入字符 %s %s" % (str(len(charDict)), c))
                charDict.add(c)


def loadBaseChars(charDict):
    filePath = os.path.join(dir, "config/Fonts_CN.txt")
    try:
        # 以utf-8编码打开文件，并处理可能存在的BOM
        with open(filePath, 'r', encoding='utf-8-sig') as f:
            lines = f.readlines()
            for line in lines:
                seg_char(line, charDict)
    except UnicodeDecodeError:
        # 如果UTF-8解码失败，尝试其他编码
        with open(filePath, 'r', encoding='gbk') as f:
            lines = f.readlines()
            for line in lines:
                seg_char(line, charDict)
    print("基础字符加载完成")

def grepFromExcelFile(filePath, charDict):
    # try:
    workbook = xl.load_workbook(filePath)
    sheetNames = workbook.get_sheet_names()
    for name in sheetNames:
        if name in skipFiles:
            print('过滤了 %s' % name);
            continue;

        sheet = workbook.get_sheet_by_name(name)

        #遍历所有行，将excel转化为json对象
        j = 2
        jName = sheet.cell(2, j).value;
        while j < 50 or jName is not None:
            # Python 3中所有字符串都是Unicode，只需要检查是否为str类型
            if isinstance(jName, str):
                if 'name' in jName or 'desc' in jName or True:
                    # print("开始处理", jName, name)
                    for i in range(5, sheet.max_row + 1):
                        if sheet.cell(i, 1).value == "END":
                            break
                        else:
                            value = sheet.cell(i, j).value
                            if isinstance(value, str):
                                seg_char(value, charDict)

            j = j + 1
            jName = sheet.cell(2, j).value
    
    # except Exception, e:
    #     print 'excel表格读取失败：%s %s' % (filePath, e)


def grepFromFile(filePath, needMatchMultLine, charDict):
    fName = path.basename(filePath);
    if fName in skipFiles:
        # print '过滤了', fName;
        return 

    bOK = False;
    for includeFileEnd in includeFilesRegular:
        if fName.endswith(includeFileEnd):
            bOK = True;
            break;

    if not bOK:
        # print '过滤了', fName;
        return ;


    isUTF8 = checkIsUTF8(filePath)
    rMULTI_DOT = re.compile(r'`([^`]*)`', flags = re.DOTALL)         # `内整段非`，多行
    DOT_ONE = ['\'', '"']
    IGNORE_DOT_ONE = ['\\"', '\\\'']
    rIGNORE_FUNC = re.compile(r'cc\.log\(.*\)|cc\.warn\(.*\)|cc\.error\(.*\)|this\.log\(.*\)|console\.log\(.*\)|console\.error\(.*\)|console\.warn\(.*\)')

    s = ""
    fileName = path.basename(filePath)
    # log('当前操作' + fileName)
    with io.open(filePath, 'rt', encoding= 'utf-8') as f:
        s = f.read()
        if not isUTF8:
            # print '该文件不是utf8', filePath
            s = s.decode('utf-8')

     # 截取注释部分
    results = rIGNORE_FUNC.findall(s)
    for result in results:
        # print result
        s = s.replace(result, "")

    # re提取部分，整个文本作为字符串
    if needMatchMultLine:
        # 首先提取多行的``
        results = rMULTI_DOT.findall(s)
        for result in results:
            seg_char(result, charDict);

    # 语义分析部分，逐行分析
    lines = s.split('\n')
    for line in lines:
        stack = []
        isInDOT = False
        curDotFlag = None           #当前引号类型，比如单引号只接受单引号结尾
        for index, char in enumerate(line):
            if isInDOT:
                stack.append(char)

            if char in DOT_ONE and (not curDotFlag or curDotFlag == char):
                if isInDOT:
                    # 检查当前stack的末尾是否特殊组合
                    if ''.join(stack[-2:]) in IGNORE_DOT_ONE:
                        # print '当前dot不作为结束'
                        continue
                    # 已经在引号内，去除最后一次的引号，结算
                    content = ''.join(stack[:-1])
                    stack = []
                    seg_char(content, charDict);
                else:
                    isInDOT = not isInDOT
                    curDotFlag = char

    # print "打印字符", ret


def workFlowForGrep():
    chars = set([])      # 完整的CNKey

    loadBaseChars(chars);

    # # 处理配置表  config里的xlsx
    # for fpathe, dirs, fs in os.walk(os.path.join(dir, "config")): 
    #     for f in fs:
    #         if f.endswith('.xlsx') and f[0] != '~':
    #             print(f)
    #             grepFromExcelFile(path.join(fpathe, f), chars);

    # #处理代码
    # scriptPath = os.path.join(dir, "Script")
    # for fpathe, dirs, fs in os.walk(scriptPath):
    #     for f in fs:
    #         # print path.join(fpathe, f);
    #         grepFromFile(path.join(fpathe, f), False, chars);

    # #处理UI
    # uiPath = os.path.join(dir, "UI")
    # for fpathe, dirs, fs in os.walk(uiPath):
    #     for f in fs:
    #         grepFromFile(path.join(fpathe, f), False, chars);

    charStr = ""
    for c in chars:
        # 处理Python 2和Python 3的兼容性
        if sys.version_info[0] >= 3:
            charStr += c  # Python 3直接拼接字符串
        else:
            charStr += c.encode('utf-8')  # Python 2需要编码

    print("字符集: %s" % charStr)
    return charStr


def loadStream(fileName, *args):
    ret = ""
    try:
        # 以utf-8编码打开文件，并处理可能存在的BOM
        with open(fileName, 'r', encoding='utf-8-sig') as f:
            lines = f.readlines()
    except UnicodeDecodeError:
        # 如果UTF-8解码失败，尝试其他编码
        with open(fileName, 'r', encoding='gbk') as f:
            lines = f.readlines()
    for line in lines:
        ret += line;
    f.close()
    if len(ret) > 0:
        index = 0
        for arg in args:
            # print arg, type(arg)
            ret = ret.replace("${%s}" % index, str(arg));
            index=index+1

        return ret;
    else:
        raise Exception("filename, file data error")


def makeFont(name, makeCofnig):

    #获取字符集
    content = ""
    if makeCofnig['way'] == 'glup':
        content = workFlowForGrep();
    else:
        content = makeCofnig['content']
    makeCofnig['content'] = content

    #生成html
    outHtml = path.join(dir, "tmp", name +'.html');
    if os.path.exists(outHtml):
        os.remove(outHtml);

    htmlStr = loadStream(tempHtml, name, content);
    fo = open(outHtml, "w", encoding='utf-8')
    fo.write(htmlStr)
    fo.close();

    #生成css
    outCss = path.join(dir, "tmp", name + '.css');
    if os.path.exists(outCss):
        os.remove(outCss);

    cssStr = loadStream(tempCss, name, makeCofnig['font-family']);
    fo = open(outCss, "w", encoding='utf-8')
    fo.write(cssStr)
    fo.close()

    print("生成结束 %s" % name)

def generate(sourceFont, fontSet, outFile):
    #TTF源文
    font = subset.load_font(sourceFont, subset.Options())  # NOTE 这个方式文字一样 md5没变化

    #挑出需要的字符
    subsetter = subset.Subsetter()
    subsetter.populate(text=fontSet)
    subsetter.subset(font)

    # 生成输出文件
    subset.save_font(font, outFile, subset.Options())

    font.close()

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='字符提取')
    parser.add_argument('--brance', action="store", dest="brance", help='分支')
    parser.add_argument('--configDir', action="store", dest="configDir", help='配置分支')
    parser.add_argument('--rebuid', action="store_true", dest="rebuid", help='rebuid')
    parser.add_argument('--dryRun', action="store_true", dest="dryRun", help='dryRun')
    parser.add_argument('--svnop', action="store_true", dest="svnop", help='svnop')
    # start = tm.clock()
    isCopyFile = True
    configPath = os.path.join(dir, "config\\config.json");
    with open(configPath, 'r', encoding='utf-8') as f:
        jsonData = json.loads(f.read())
        #是否重构
        if True:
            tmpDir = os.path.join(dir, "tmp");
            if os.path.exists(tmpDir):
                shutil.rmtree(tmpDir) #删除

            os.mkdir(tmpDir); #创建

            for baseName in jsonData:
                #生成配置文件
                makeFont(baseName, jsonData[baseName]);

            with open(configPath,'w', encoding='utf-8') as fForW:
                json.dump(jsonData, fForW, indent=4, sort_keys=True, ensure_ascii=False)
        
        #批量生成字体, 不用font-spider了，对某些字库支持不好，处理不了，改用Python的fontTools
        # os.system("font-spider ./tmp/*.html");
        for baseName in jsonData:
            srcFontFile = os.path.join(dir, "源文件", baseName);
            destFontFile = os.path.join(dir, "tmp", baseName);
            generate(srcFontFile, jsonData[baseName]['content'], destFontFile)


        # 输出md5
        for baseName in jsonData:
            fontPath = os.path.join(dir, "tmp", baseName);
            md5 = hashlib.md5(open(fontPath, 'rb').read()).hexdigest()
            print(f"{baseName}: {md5}")
        if isCopyFile:
            for baseName in jsonData:
                makeCofnig = jsonData[baseName];

                #更新文件
                fontPath = os.path.join(dir, "tmp", baseName);
                destFile = os.path.join(dir, "输出文件", baseName)

                if os.path.exists(fontPath):
                    if os.path.exists(destFile):
                        os.remove(destFile);

                #拷贝
                # print "替换字体", destFile;
                shutil.copyfile(fontPath, destFile);

                # #提交
                # if results.svnop:
                #     os.system('svn commit -m "update font" %s' % destFile);

    # print("All done! It takes %s second(s)" % str(tm.clock() - start))
